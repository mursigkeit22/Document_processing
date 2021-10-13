from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


workbook = load_workbook(filename="work_docs/excel_misha/Microsoft_Excel_Worksheet.xlsx")

for sheet in workbook.worksheets:
    for column in sheet.columns:
        sheet.column_dimensions[get_column_letter(column[0].column)].hidden = True

workbook.save("work_docs/excel_misha/done/Microsoft_Excel_Worksheet.xlsx")
for i in range(1, 33):
    workbook = load_workbook(filename=f"work_docs/excel_misha/Microsoft_Excel_Worksheet{i}.xlsx")
    for sheet in workbook.worksheets:
        for column in sheet.columns:
            sheet.column_dimensions[get_column_letter(column[0].column)].hidden = True
    workbook.save(f"work_docs/excel_misha/done/Microsoft_Excel_Worksheet{i}.xlsx")
