from openpyxl import load_workbook

workbook = load_workbook(filename="./excel_docs/_6_Downstream General TB_CONCAT.xlsx")
for sheet in workbook.worksheets:
    for row in sheet.iter_rows():
        if row[0].value.strip().startswith("|"):
            print(row[0].value)
