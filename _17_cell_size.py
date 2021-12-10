from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import math

workbook = load_workbook(filename="./work_docs/reseda_done.xlsx", data_only=True)
ws = workbook["ТКП с подписью"]

"""
for Arial 11
"""
max_width = 40
max_letters_in_line = 40
max_height = 80
max_lines = 5
ws.column_dimensions['A'].width = 40
ws.row_dimensions[5].height = 40
col_range = ws[ws.min_column: ws.max_column]
# print(ws.min_column)
# print(ws.max_column)
# print(col_range)
for row_cells in ws.iter_rows():
    max_length = 0
    max_cell = None
    for cell in row_cells:
        if cell.value:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
                max_cell = cell
    if max_cell:
        split_text = max_cell.value.split("\n")
        longest_line = max(split_text, key=len)

    print(max_length, max_cell, max_cell.value if max_cell else None)

    print("-----------------------------")



# workbook.save("./work_docs/reseda_done2.xlsx")