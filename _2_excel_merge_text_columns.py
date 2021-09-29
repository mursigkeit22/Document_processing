from openpyxl import load_workbook

workbook_trans = load_workbook(filename="work_docs/Trans.xlsm")
workbook_orig = load_workbook(filename="work_docs/Orig.xlsm", read_only=False, keep_vba=True)

sheet_trans = workbook_trans["_Востребованность_"]
sheet_orig = workbook_orig["Спецификация"]


def check_same_rows_number(sheet1, sheet2, ):
    if sheet1.max_row != sheet2.max_row:
        raise Exception("Different number of rows in documents.")


check_same_rows_number(sheet_trans, sheet_orig)

trans_text_list = []
for cell in sheet_trans["C"]:
    trans_text_list.append(cell.value)

for i in range(1, 168):
    try:
        old_cell_value = sheet_orig.cell(row=i, column=3).value
        new_cell_value = trans_text_list[i - 1] + "\n" + "\n" + old_cell_value
        sheet_orig.cell(row=i, column=3).value = new_cell_value
    except TypeError:  # for empty rows
        pass

workbook_orig.save("./work_docs/new_orig.xlsm")
