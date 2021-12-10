from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
import math

"""
len(next(ws.iter_rows())) - число колонок в листе
"""

workbook = load_workbook(filename="./work_docs/reseda_done.xlsx", data_only=True)
ws = workbook["ТКП с подписью"]


col_width = []
for i in range(len(next(ws.iter_rows()))):  # итерируемся по всем колонкам
    col_letter = get_column_letter(i + 1)

    minimum_width = 20
    current_width = ws.column_dimensions[col_letter].width  # получаем текущую ширину колонки
    if not current_width or current_width < minimum_width:
        ws.column_dimensions[col_letter].width = minimum_width  # увеличиваем ширину колонки, если она меньше минимума

    col_width.append(ws.column_dimensions[col_letter].width)  # добавляем ширину колонки в аррей

for i, row in enumerate(ws): # i - номер ряда, начиная с нуля; row - все ячейки в этом ряду
    default_height = 12.5  # Corresponding to font size 12

    multiples_of_font_size = [default_height]
    for j, cell in enumerate(row): # j - номер колонки, начиная с нуля
        wrap_text = True
        vertical = "top"
        if cell.value is not None:
            mul = 0
            for v in str(cell.value).split('\n'):
                print(v)
                mul += math.ceil(len(v) / col_width[j]) * cell.font.size
            print(mul)
#             if mul > 0:
#                 multiples_of_font_size.append(mul)
# #
#         cell.alignment = Alignment(wrap_text=wrap_text, vertical=vertical)
#
#     original_height = ws.row_dimensions[i + 1].height
#     if original_height is None:
#         original_height = default_height
#
#     new_height = max(multiples_of_font_size)
#     if original_height < new_height:
#         ws.row_dimensions[i + 1].height = new_height
#
# workbook.save("./work_docs/reseda_done2.xlsx")