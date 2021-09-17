from openpyxl import load_workbook
import re

"""
сначала проверяем на пустые строки
"""

workbook = load_workbook(filename="./excel_docs/test_alpha.xlsx")
work_sheet = workbook["Лист1"]


def delete_raw_with_empty_cell(work_sheet, column_num): # TODO: check if it works with first raw!
    """ Column numbers are 1, 2, 3 ... Not 0, 1, 2, 3!!!"""
    for rowNum in range(work_sheet.max_row, 0, -1):
        print(rowNum)
        # print(work_sheet.cell(row=rowNum, column=2))
        if work_sheet.cell(row=rowNum, column=column_num).value is None:
            # print("hop")
            work_sheet.delete_rows(rowNum)

delete_raw_with_empty_cell(work_sheet, 1)