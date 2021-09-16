from openpyxl import load_workbook
import re

workbook = load_workbook(filename="./documents_excel/test_python.xlsx")
sheet_to_focus = 'Лист1'

workbook.active = 0
work_sheet = workbook.active

for rowNum in range(work_sheet.max_row, 1, -1):
    print(work_sheet.cell(row=rowNum, column=2))
    if work_sheet.cell(row=rowNum, column=2).value is None:
        print("hop")
        work_sheet.delete_rows(rowNum)


