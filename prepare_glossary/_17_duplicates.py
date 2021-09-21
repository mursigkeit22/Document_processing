from openpyxl import load_workbook, Workbook

from prepare_glossary.helper_functions import delete_raw_with_empty_cell

""" 
колонки в доках должны быть English, Russian, Notes1, Notes2

    [
    [0 English, 1 Russian, 2 Notes1, 3 Notes2],
    [0 English, 1 Russian, 2 Notes1, 3 Notes2],
    ...
    ]
    В готовой ТБ нужно будет убрать карточки English-Russian
"""

workbook = load_workbook(filename="./excel_docs/KTK_Prepared_full.xlsx")

#  вытаскиваем данные из всех листов, записываем в список.
#  при этом удаляем пустые строки и проверяем,
#  чтобы не осталось Notes, в которых только пробелы

data = []

for sheet in workbook.worksheets:

    delete_raw_with_empty_cell(sheet, 1)
    delete_raw_with_empty_cell(sheet, 2)
    for row in sheet.iter_rows():
        temp_list = [cell.value for cell in row][:4]
        if temp_list[2]:
            if len(temp_list[2].strip()) < 1:
                temp_list[2] = None
        if temp_list[3]:
            if len(temp_list[3].strip()) < 1:
                temp_list[3] = None
        data.append(temp_list)

print("finished iteration.")
print("Starting data length: ", len(data))
# чистим графу с английским от пробельных знаков, чтобы сортировалось правильно
for el in data:
    el[0] = el[0].strip()

# сортируем только по элементу с английским, не учитывая регистр
print("Full_sort_started")
data.sort(key=lambda x: x[0].casefold())

# дальше работаем только со строками, в которых нет Notes
# (чтобы комментарии относились только к тем строкам, к которым относились изначально)

data_with_notes = [x for x in data if x[2] or x[3]]
print("data_with_notes: ", len(data_with_notes))
data_without_notes = [x for x in data if not x[2] and not x[3]]
print("data_without_notes: ", len(data_without_notes))


# соединяем дубликаты
i = 0
while i < len(data_without_notes) - 1:
    current_row = data_without_notes[i]
    next_row = data_without_notes[i + 1]
    current_temp_english = current_row[0].split("|")
    current_temp_english_lower = [x.lower() for x in current_temp_english]
    next_temp_english = next_row[0].split("|")
    next_temp_english_lower = [x.lower() for x in next_temp_english]

    for el in current_temp_english_lower:
        if el in next_temp_english_lower:
            data_without_notes[i + 1] = [current_row[0] + "|" + next_row[0],
                                         current_row[1] + "|" + next_row[1],
                                         None, None]
            data_without_notes[i] = [None]

    i += 1
# удаляем пустые строки, образовавшиеся на месте дубликатов
data_without_notes = [x for x in data_without_notes if len(x) > 1]

# чистим от повторяющихся элементов внутри языков
ready_data = []
i = 0
while i < len(data_without_notes):
    split_english = data_without_notes[i][0].split("|")
    split_russian = data_without_notes[i][1].split("|")
    split_english_ready = [x.strip().lower() for x in split_english]
    split_russian_ready = [x.strip().lower() for x in split_russian]
    english_set = set(split_english_ready)
    russian_set = set(split_russian_ready)

    new_english = []

    english_index = []
    for unique_el in english_set:
        english_index.append(split_english_ready.index(unique_el))
    new_english = [split_english[x] for x in english_index]
    new_english_string = " | ".join(new_english)

    new_russian = []
    russian_index = []
    for unique_el in russian_set:
        russian_index.append(split_russian_ready.index(unique_el))
    new_russian = [split_russian[x] for x in russian_index]
    new_russian_string = " | ".join(new_russian)

    ready_data.append([new_english_string, new_russian_string, None, None])

    i += 1

ready_data = ready_data + data_with_notes
print("len(ready_data): ", len(ready_data))

# записываем в файл
new_wb = Workbook()

ws = new_wb.active
for row in ready_data:
    ws.append(row)

new_wb.save("./ready_excel_docs/done_KTK_Prepared_full.xlsx")
