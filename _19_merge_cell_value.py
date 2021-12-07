from openpyxl import load_workbook
from openpyxl.styles import Alignment


workbook_orig_rus = load_workbook(filename="./work_docs/ориг_русс.xlsx", data_only=True)
workbook_trans_eng = load_workbook(filename="./work_docs/перевод_англ.xlsx", data_only=True)

sheet_orig_rus = workbook_orig_rus["ТКП с подписью"]
sheet_trans_eng = workbook_trans_eng["TCP with signature"]


for row_cells in sheet_orig_rus.iter_rows():
    for cell in row_cells:
        if cell.value is not None and isinstance(cell.value, str):
            old_cell_value = cell.value
            print("rus", old_cell_value)
            cell_value_english = sheet_trans_eng[cell.coordinate].value
            new_cell_value = old_cell_value + " /" + "\n" + cell_value_english
            cell.value = new_cell_value
            sheet_orig_rus[cell.coordinate].alignment = Alignment(wrapText=True)
            
workbook_orig_rus.save("./work_docs/reseda_done.xlsx")